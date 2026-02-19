#!/usr/bin/env python3
"""
Générateur Excel — Grille d'Objectifs Annuels avec calcul automatique de prime.

Usage:
    python generate_excel.py --config config.json --output Objectifs_2026_NOM.xlsx

Le fichier config.json doit contenir :
{
  "collaborateur": "Prénom NOM",
  "poste": "Titre du poste",
  "entite": "Deskea / Tersea",
  "annee": 2026,
  "manager": "Nom du manager",
  "bareme": [                          // optionnel, défaut fourni
    {"seuil": "<40%", "niveau": "Insatisfaisant", "prime": "0%", "couleur": "F44336"},
    ...
  ],
  "axes": [
    {
      "id": "AXE 1",
      "label": "STRATÉGIE PRODUIT (20%)",
      "couleur": "0984E3",              // hex sans #
      "objectifs": [
        {
          "num": 1,
          "titre": "Intitulé de l'objectif",
          "axe_strategique": "Innovation & R&D",
          "type": "Quantitatif",
          "poids": 0.10,
          "kpi": "Description KPI + flags ⚠️",
          "cible_quanti": 4,            // ou "" si qualitatif
          "cible_quali": "Description cible qualitative détaillée",
          "seuil": 3,                   // ou "" si qualitatif
          "mode_calcul": "Cumul",       // Cumul | Pourcentage | Moyenne | Autre
          "periode": "Trimestriel",
          "echeance": "31/12/2026",
          "ind_coll": "Individuel",
          "plan_action": "1) Étape 1 2) Étape 2 3) Étape 3"
        }
      ]
    }
  ]
}
"""

import json
import sys
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


# ── Defaults ──────────────────────────────────────────────────────────────────

DEFAULT_BAREME = [
    {"seuil": "< 40%",    "niveau": "Insatisfaisant",             "prime": "0% prime",   "couleur": "F44336"},
    {"seuil": "40% - 59%","niveau": "Partiellement satisfaisant", "prime": "50% prime",  "couleur": "FF9800"},
    {"seuil": "60% - 79%","niveau": "Satisfaisant",               "prime": "100% prime", "couleur": "4CAF50"},
    {"seuil": "80% - 89%","niveau": "Très satisfaisant",          "prime": "120% prime", "couleur": "2196F3"},
    {"seuil": "≥ 90%",    "niveau": "Exceptionnel",               "prime": "150% prime", "couleur": "9C27B0"},
]

AXE_COLORS_DEFAULT = ["0984E3", "00B894", "E17055", "6C5CE7", "E84393"]


# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Arial")
HEADER_FILL = PatternFill("solid", fgColor="2D3436")
EVAL_HEADER_FILL = PatternFill("solid", fgColor="1B5E20")
AXE_FONT = Font(bold=True, color="FFFFFF", size=10, name="Arial")
DATA_FONT = Font(size=9, name="Arial")
INPUT_FONT = Font(size=10, name="Arial", bold=True, color="0000FF")
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
RESULT_FONT = Font(size=10, name="Arial", bold=True)
RESULT_FILL = PatternFill("solid", fgColor="E8F5E9")
TOTAL_FONT = Font(size=12, name="Arial", bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="2D3436")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"), right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"), bottom=Side(style="thin", color="D5D8DC"),
)
THICK_BORDER = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"), bottom=Side(style="medium"),
)

HEADERS = [
    "#", "Objectif", "Axe stratégique", "Type objectif", "Poids (%)",
    "KPI (+ mise en place si inexistant)", "Cible quanti.", "Cible qualitative",
    "Seuil minimum", "Mode calcul", "Période éval.", "Échéance", "Ind./Coll.",
    "Plan d'action détaillé", "",
    "RÉSULTAT ATTEINT (saisie)", "COMMENTAIRE ÉVALUATION", "NOTE /5 (saisie)",
    "NOTE PONDÉRÉE", "% ATTEINTE",
]

COL_WIDTHS = {
    "A": 4, "B": 42, "C": 17, "D": 11, "E": 7, "F": 36, "G": 9, "H": 45,
    "I": 9, "J": 11, "K": 11, "L": 11, "M": 10, "N": 42, "O": 3,
    "P": 22, "Q": 28, "R": 12, "S": 13, "T": 12,
}


def generate(config: dict, output_path: str):
    collaborateur = config["collaborateur"]
    annee = config.get("annee", 2026)
    bareme = config.get("bareme", DEFAULT_BAREME)
    axes = config["axes"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Objectifs {annee} - {collaborateur.split()[-1]}"

    # ── Headers ───────────────────────────────────────────────────────────
    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL if col_idx <= 14 else (EVAL_HEADER_FILL if col_idx >= 16 else PatternFill())
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER

    # ── Data ──────────────────────────────────────────────────────────────
    row = 2
    data_rows = []

    for axe_idx, axe in enumerate(axes):
        color = axe.get("couleur", AXE_COLORS_DEFAULT[axe_idx % len(AXE_COLORS_DEFAULT)])
        fill = PatternFill("solid", fgColor=color)

        # Axe header row
        for col_idx in range(1, 21):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = axe["id"] if col_idx == 1 else (axe["label"] if col_idx == 2 else "")
            cell.font = AXE_FONT
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            cell.border = THIN_BORDER
        row += 1

        # Objective rows
        for obj in axe["objectifs"]:
            values = [
                obj["num"], obj["titre"], obj["axe_strategique"], obj["type"],
                obj["poids"], obj["kpi"], obj.get("cible_quanti", ""),
                obj.get("cible_quali", ""), obj.get("seuil", ""),
                obj.get("mode_calcul", ""), obj.get("periode", ""),
                obj.get("echeance", ""), obj.get("ind_coll", ""),
                obj.get("plan_action", ""),
                "",  # col O spacer
                "", "", "",  # P, Q, R — input
                f'=IF(R{row}="","",R{row}*E{row})',   # S — note pondérée
                f'=IF(R{row}="","",R{row}/5)',          # T — % atteinte
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = val
                cell.border = THIN_BORDER
                if col_idx <= 14:
                    cell.font = DATA_FONT
                    cell.alignment = WRAP
                    if col_idx == 5:
                        cell.number_format = '0%'
                elif col_idx in (16, 17, 18):
                    cell.font = INPUT_FONT
                    cell.fill = INPUT_FILL
                    cell.alignment = WRAP_CENTER if col_idx == 17 else CENTER
                elif col_idx in (19, 20):
                    cell.font = RESULT_FONT
                    cell.fill = RESULT_FILL
                    cell.alignment = CENTER
                    cell.number_format = '0.00' if col_idx == 19 else '0%'

            data_rows.append(row)
            row += 1

    # ── Data validation for Note /5 ───────────────────────────────────────
    dv = DataValidation(
        type="decimal", operator="between", formula1=0, formula2=5,
        allow_blank=True, showErrorMessage=True,
        errorTitle="Note invalide", error="La note doit être entre 0 et 5.",
    )
    dv.promptTitle = "Note /5"
    dv.prompt = "Saisir une note entre 0 et 5"
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    for r in data_rows:
        dv.add(ws.cell(row=r, column=18))

    # ── Summary block ─────────────────────────────────────────────────────
    row += 1
    r_cells = ",".join([f"R{r}" for r in data_rows])
    s_cells = ",".join([f"S{r}" for r in data_rows])

    # Header
    for c in range(16, 21):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THICK_BORDER
        cell.alignment = CENTER
    ws.cell(row=row, column=16).value = "SYNTHÈSE"
    ws.merge_cells(start_row=row, start_column=16, end_row=row, end_column=17)
    ws.cell(row=row, column=18).value = "Moy."
    ws.cell(row=row, column=19).value = "Pondérée"
    ws.cell(row=row, column=20).value = "% Prime"
    row += 1

    # Totals
    for c in range(16, 21):
        cell = ws.cell(row=row, column=c)
        cell.border = THICK_BORDER
        cell.alignment = CENTER
        cell.font = Font(size=11, name="Arial", bold=True)
        cell.fill = RESULT_FILL

    ws.cell(row=row, column=16).value = "NOTE GLOBALE"
    ws.merge_cells(start_row=row, start_column=16, end_row=row, end_column=17)
    ws.cell(row=row, column=18).value = f'=IF(COUNT({r_cells})=0,"",AVERAGE({r_cells}))'
    ws.cell(row=row, column=18).number_format = '0.00'
    ws.cell(row=row, column=19).value = f'=IF(COUNT({r_cells})=0,"",SUM({s_cells}))'
    ws.cell(row=row, column=19).number_format = '0.00'
    summary_row = row
    ws.cell(row=row, column=20).value = f'=IF(S{row}="","",S{row}/5)'
    ws.cell(row=row, column=20).number_format = '0.0%'
    ws.cell(row=row, column=20).font = Font(size=14, name="Arial", bold=True, color="1B5E20")
    ws.cell(row=row, column=20).fill = PatternFill("solid", fgColor="C8E6C9")
    row += 1

    # Explanation
    ws.cell(row=row, column=16).value = (
        "Lecture : 100% = tous les objectifs atteints au niveau 5/5 (exceptionnel). "
        "Le % d'atteinte pondéré sert de base au calcul de la prime variable."
    )
    ws.merge_cells(start_row=row, start_column=16, end_row=row, end_column=20)
    ws.cell(row=row, column=16).font = Font(size=8, name="Arial", italic=True, color="666666")
    ws.cell(row=row, column=16).alignment = WRAP
    row += 2

    # Barème
    ws.cell(row=row, column=16).value = "BARÈME INDICATIF"
    ws.merge_cells(start_row=row, start_column=16, end_row=row, end_column=20)
    for c in range(16, 21):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="37474F")
        ws.cell(row=row, column=c).font = Font(size=10, name="Arial", bold=True, color="FFFFFF")
        ws.cell(row=row, column=c).alignment = CENTER
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

    for item in bareme:
        ws.cell(row=row, column=16).value = item["seuil"]
        ws.cell(row=row, column=16).font = Font(size=9, name="Arial", bold=True)
        ws.merge_cells(start_row=row, start_column=17, end_row=row, end_column=18)
        ws.cell(row=row, column=17).value = item["niveau"]
        ws.cell(row=row, column=17).font = Font(size=9, name="Arial")
        ws.merge_cells(start_row=row, start_column=19, end_row=row, end_column=20)
        ws.cell(row=row, column=19).value = item["prime"]
        ws.cell(row=row, column=19).font = Font(size=9, name="Arial", bold=True, color=item["couleur"])
        for c in range(16, 21):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = CENTER
        row += 1

    # ── Column widths & freeze ────────────────────────────────────────────
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    for r in range(2, row):
        v = ws.cell(row=r, column=1).value
        if v and str(v).startswith("AXE"):
            ws.row_dimensions[r].height = 25
        elif v and str(v).isdigit():
            ws.row_dimensions[r].height = 85
        else:
            ws.row_dimensions[r].height = 22

    wb.save(output_path)
    return {"rows": len(data_rows), "file": output_path}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate HR Objectives Excel")
    parser.add_argument("--config", required=True, help="Path to config JSON")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    with open(args.config) as f:
        cfg = json.load(f)

    result = generate(cfg, args.output)
    print(json.dumps(result, indent=2))
